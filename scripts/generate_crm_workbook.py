#!/usr/bin/env python3
"""Generate the close-day CRM workbook template and optional CSV seed files."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]

SHEETS = {
    "Accounts": [
        "Account Name",
        "Relationship Type",
        "Stage",
        "Status",
        "Priority",
        "Owner",
        "Last Touch",
        "Next Follow-up",
        "Next Step",
        "Source Evidence",
        "Canonical Page URL",
    ],
    "Contacts": [
        "Name",
        "Account",
        "Role/Title",
        "Email",
        "Relationship Role",
        "Last Touch",
        "Follow-up Flag",
        "Notes",
    ],
    "Interactions": [
        "Date",
        "Channel",
        "Account",
        "Contacts",
        "Subject",
        "Summary",
        "Action Extracted",
        "Source Link",
    ],
    "FollowUps": [
        "Due Date",
        "Account",
        "Contact",
        "Ask/Task",
        "Owner",
        "Status",
        "Source Interaction",
    ],
}

LISTS = {
    "Stages": [
        "Lead",
        "Pursuit",
        "Pilot",
        "Award Onboarding",
        "Active Program",
        "Active Collaboration",
        "Partner Pipeline",
        "Closed",
    ],
    "Statuses": [
        "Active",
        "Waiting",
        "At Risk",
        "On Hold",
        "Won",
        "Lost",
        "Closed",
    ],
    "Relationship Types": [
        "Customer/Program",
        "Opportunity",
        "Research/Customer Collaboration",
        "Award/Program",
        "Partner/Ecosystem",
        "Partner-led Pipeline",
    ],
    "Priorities": [
        "P1 - Must",
        "P2 - Should",
        "P3 - Could",
        "P4 - Later",
    ],
    "Channels": [
        "Gmail",
        "Outlook",
        "Slack",
        "Teams",
        "Granola",
        "Meeting",
        "Manual",
    ],
    "Follow-up Statuses": [
        "Open",
        "Waiting",
        "Done",
        "Canceled",
    ],
    "Relationship Roles": [
        "Customer",
        "Sponsor",
        "Partner",
        "Program Manager",
        "Technical",
        "Contracting",
        "Internal",
    ],
    "Follow-up Flags": [
        "Yes",
        "No",
    ],
    "Owners": [
        "Owner",
        "Delegate",
        "Team",
    ],
}

VALIDATIONS = {
    ("Accounts", "B"): "Lists!$C$2:$C$7",
    ("Accounts", "C"): "Lists!$A$2:$A$9",
    ("Accounts", "D"): "Lists!$B$2:$B$8",
    ("Accounts", "E"): "Lists!$D$2:$D$5",
    ("Accounts", "F"): "Lists!$I$2:$I$4",
    ("Contacts", "E"): "Lists!$G$2:$G$8",
    ("Contacts", "G"): "Lists!$H$2:$H$3",
    ("Interactions", "B"): "Lists!$E$2:$E$8",
    ("FollowUps", "E"): "Lists!$I$2:$I$4",
    ("FollowUps", "F"): "Lists!$F$2:$F$5",
}

DATE_COLUMNS = {
    ("Accounts", "G"),
    ("Accounts", "H"),
    ("Contacts", "F"),
    ("Interactions", "A"),
    ("FollowUps", "A"),
}

WIDTHS = {
    "Accounts": [26, 24, 20, 16, 14, 14, 14, 16, 34, 42, 34],
    "Contacts": [24, 26, 24, 30, 20, 14, 16, 42],
    "Interactions": [14, 14, 26, 30, 34, 46, 34, 42],
    "FollowUps": [14, 26, 24, 42, 14, 14, 42],
}


def resolve(path_text: str) -> Path:
    path = Path(path_text)
    return path if path.is_absolute() else ROOT / path


def style_sheet(ws, headers: list[str], widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin_gray)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.append(headers)
    ws.append(["" for _ in headers])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2, max_row=200, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    table_ref = f"A1:{chr(64 + len(headers))}2"
    table_name = f"{ws.title}Table"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_lists_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Lists")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    fill = PatternFill("solid", fgColor="44546A")
    font = Font(color="FFFFFF", bold=True)
    for col_idx, (header, values) in enumerate(LISTS.items(), start=1):
        ws.cell(1, col_idx, header)
        ws.cell(1, col_idx).fill = fill
        ws.cell(1, col_idx).font = font
        ws.cell(1, col_idx).alignment = Alignment(horizontal="center")
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)
        ws.column_dimensions[chr(64 + col_idx)].width = max(16, len(header) + 2)
    ws.freeze_panes = "A2"


def add_data_validations(wb: Workbook) -> None:
    for (sheet_name, col), formula in VALIDATIONS.items():
        ws = wb[sheet_name]
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}200")
    for sheet_name, col in DATE_COLUMNS:
        ws = wb[sheet_name]
        for row in range(2, 201):
            ws[f"{col}{row}"].number_format = "yyyy-mm-dd"


def write_csv_seed(csv_dir: Path) -> None:
    csv_dir.mkdir(parents=True, exist_ok=True)
    for sheet, headers in SHEETS.items():
        with (csv_dir / f"{sheet}.csv").open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
    with (csv_dir / "Lists.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        headers = list(LISTS)
        writer.writerow(headers)
        max_len = max(len(values) for values in LISTS.values())
        for row_idx in range(max_len):
            writer.writerow([
                LISTS[header][row_idx] if row_idx < len(LISTS[header]) else ""
                for header in headers
            ])


def create_workbook(output: Path, csv_dir: Path | None) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, headers in SHEETS.items():
        ws = wb.create_sheet(sheet)
        style_sheet(ws, headers, WIDTHS[sheet])
    add_lists_sheet(wb)
    add_data_validations(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    if csv_dir:
        write_csv_seed(csv_dir)


def verify_workbook(path: Path) -> dict:
    wb = load_workbook(path)
    result = {"sheets": wb.sheetnames, "tables": {}, "validations": {}}
    expected = set(SHEETS) | {"Lists"}
    missing = sorted(expected - set(wb.sheetnames))
    if missing:
        raise ValueError(f"missing sheets: {missing}")
    for sheet, headers in SHEETS.items():
        ws = wb[sheet]
        actual_headers = [cell.value for cell in ws[1][: len(headers)]]
        if actual_headers != headers:
            raise ValueError(f"{sheet}: header mismatch: {actual_headers}")
        result["tables"][sheet] = list(ws.tables.keys())
        result["validations"][sheet] = len(ws.data_validations.dataValidation)
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate close-day CRM workbook template.")
    parser.add_argument("--output", default="assets/crm/daily-close-crm-template.xlsx")
    parser.add_argument("--csv-dir", help="Optional CSV seed directory.")
    parser.add_argument("--verify-only", action="store_true")
    args = parser.parse_args()

    try:
        output = resolve(args.output)
        csv_dir = resolve(args.csv_dir) if args.csv_dir else None
        if not args.verify_only:
            create_workbook(output, csv_dir)
        result = verify_workbook(output)
        print(f"validated: {output}")
        print(f"sheets: {', '.join(result['sheets'])}")
        return 0
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
