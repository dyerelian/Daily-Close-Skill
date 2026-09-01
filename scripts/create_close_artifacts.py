#!/usr/bin/env python3
"""Create portable close-day Markdown/JSON artifacts from an approved payload."""

from __future__ import annotations

import argparse
import copy
import json
import re
from datetime import date
from pathlib import Path
from typing import Any

from close_day_config import (
    atomic_write_json,
    atomic_write_text,
    export_enabled,
    load_json,
    resolve_profile,
    resolved_artifact_paths,
    validate_profile,
)
from create_agenda_docx import create_docx as create_agenda_docx
from create_daily_plan_docx import create_docx as create_daily_plan_docx
from close_payload import configured_gtd_link, normalize_payload


def clean_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "-", value).strip(" .")
    return cleaned or "agenda"


def iso_date(value: object, field: str) -> str:
    try:
        return date.fromisoformat(str(value)).isoformat()
    except ValueError as exc:
        raise ValueError(f"{field} must be an ISO date (YYYY-MM-DD)") from exc


def label(item: Any, scopes: dict[str, str]) -> str:
    if isinstance(item, str):
        return item
    if not isinstance(item, dict):
        return str(item)
    text = str(item.get("text") or item.get("title") or item.get("subject") or "").strip()
    when = str(item.get("time") or item.get("start") or "").strip()
    if when and (item.get("subject") or item.get("title")):
        text = f"{when} — {text}"
    location = str(item.get("location") or "").strip()
    if location:
        text = f"{text} ({location})"
    scope_id = item.get("scope_id")
    scope_name = scopes.get(scope_id, scope_id) if scope_id else None
    return f"[{scope_name}] {text}" if scope_name else text


def bullets(values: Any, scopes: dict[str, str]) -> list[str]:
    return [f"- {label(value, scopes)}" for value in (values or []) if label(value, scopes).strip()]


def section(lines: list[str], heading: str, values: Any, scopes: dict[str, str]) -> None:
    rendered = bullets(values, scopes)
    if not rendered:
        return
    lines.extend([f"## {heading}", "", *rendered, ""])


def takeaways(lines: list[str], payload: dict, scopes: dict[str, str]) -> None:
    value = payload.get("takeaways") or {}
    well = value.get("well") or []
    improve = value.get("improve") or []
    if not well and not improve:
        return
    lines.extend(["## Daily Takeaways", ""])
    if well:
        lines.append("**Did well**")
        lines.append("")
        lines.extend(bullets(well, scopes))
        lines.append("")
    if improve:
        lines.append("**To improve next time**")
        lines.append("")
        lines.extend(bullets(improve, scopes))
        lines.append("")


def crm_review_summary(lines: list[str], payload: dict, scopes: dict[str, str]) -> None:
    review = payload.get("crm_review") or {}
    if not review:
        return
    lines.extend(["## CRM Review", ""])
    status = str(review.get("status") or "unknown").replace("_", " ").title()
    lines.append(f"- Status: {status}")
    handler = str(review.get("handler_skill") or "").strip()
    if handler:
        lines.append(f"- Handler: {handler}")
    window = review.get("window") or {}
    if window.get("start") and window.get("end"):
        lines.append(f"- Review window: {window['start']} through {window['end']}")
    counts = review.get("counts") or {}
    if counts:
        rendered_counts = ", ".join(
            f"{key.replace('_', ' ')} {value}" for key, value in sorted(counts.items())
        )
        lines.append(f"- Changes: {rendered_counts}")
    lines.append("")
    section(lines, "CRM updates", review.get("summary_items"), scopes)
    section(lines, "CRM review flags", review.get("review_flags"), scopes)
    gaps = review.get("gaps") or []
    if gaps:
        lines.extend(["### CRM coverage gaps", ""])
        lines.extend(f"- {str(gap)}" for gap in gaps)
        lines.append("")


def plan_reflections(lines: list[str], payload: dict, scopes: dict[str, str]) -> None:
    value = payload.get("takeaways") or {}
    well = value.get("well") or []
    improve = value.get("improve") or []
    required = int(value.get("required_items") or max(len(well), len(improve), 0))
    noun = "thing" if required == 1 else "things"
    if well:
        lines.extend([f"## Yesterday — {required} {noun} I did well", ""])
        lines.extend(f"{index}. {label(item, scopes)}" for index, item in enumerate(well, 1))
        lines.append("")
    if improve:
        lines.extend([f"## Today — {required} {noun} I can improve", ""])
        lines.extend(f"{index}. {label(item, scopes)}" for index, item in enumerate(improve, 1))
        lines.append("")


def validate_required_takeaways(payload: dict, profile: dict) -> None:
    config = ((profile.get("features") or {}).get("daily_takeaways") or {})
    if not config.get("enabled", False):
        return
    required = int(config.get("required_items") or 0)
    if required <= 0 or config.get("incomplete_policy", "allow_partial") != "ask_until_complete":
        return
    takeaways_value = payload.get("takeaways") or {}
    missing = []
    for key, label_text in (("well", "things done well"), ("improve", "things to improve")):
        values = [item for item in (takeaways_value.get(key) or []) if label(item, {}).strip()]
        if len(values) != required:
            missing.append(f"{label_text}: expected exactly {required}, found {len(values)}")
    if missing:
        raise ValueError(
            "Daily Plan reflections are incomplete; ask the user before finalizing: "
            + "; ".join(missing)
        )


def eod_markdown(payload: dict, scopes: dict[str, str]) -> str:
    close_date = payload.get("date") or date.today().isoformat()
    lines = [f"# End-of-Day Close — {close_date}", ""]
    takeaways(lines, payload, scopes)
    crm_review_summary(lines, payload, scopes)
    sections = payload.get("sections") or {}
    section(lines, "Meeting Insights", sections.get("meeting_insights"), scopes)
    for key, heading in (
        ("accomplished", "Accomplished"),
        ("captured", "Captured"),
        ("carried", "Carried Forward"),
        ("waiting", "Waiting On"),
        ("notes", "Notes"),
    ):
        section(lines, heading, sections.get(key), scopes)
    return "\n".join(lines).rstrip() + "\n"


def plan_markdown(payload: dict, scopes: dict[str, str]) -> str:
    payload = normalize_payload(payload)
    target = payload.get("target_date") or payload.get("date") or date.today().isoformat()
    lines = [f"# Daily Plan — {target}", ""]
    plan_reflections(lines, payload, scopes)
    summary = str(payload.get("summary") or "").strip()
    if summary:
        lines.extend([summary, ""])
    sections = payload.get("sections") or {}
    section(lines, "Meeting Insights", sections.get("meeting_insights"), scopes)
    gtd_link = payload.get("gtd_link") or {}
    if gtd_link.get("url"):
        lines.extend(
            [f"[{gtd_link.get('label') or 'Open full GTD list'}]({gtd_link['url']})", ""]
        )
    for key, heading in (
        ("priorities", "Priorities"),
        ("tasks", "Tasks"),
        ("waiting", "Waiting On"),
        ("meetings", "Meetings"),
        ("people_outreach", "People Outreach"),
    ):
        section(lines, heading, sections.get(key), scopes)
    return "\n".join(lines).rstrip() + "\n"


def task_markdown(payload: dict, scopes: dict[str, str]) -> str:
    payload = normalize_payload(payload)
    target = payload.get("target_date") or payload.get("date") or date.today().isoformat()
    lines = [f"# Tasks — {target}", ""]
    tasks = ((payload.get("sections") or {}).get("tasks") or []) + ((payload.get("sections") or {}).get("carried") or [])
    lines.extend(f"- [ ] {label(item, scopes)}" for item in tasks if label(item, scopes).strip())
    return "\n".join(lines).rstrip() + "\n"


def agenda_markdown(agenda: dict, scopes: dict[str, str]) -> str:
    title = label({"text": agenda.get("title") or "Meeting Agenda", "scope_id": agenda.get("scope_id")}, scopes)
    lines = [f"# {title}", ""]
    recap = agenda.get("last_meeting_recap") or {}
    if recap:
        lines.extend(["## Last meeting recap", ""])
        summary = str(recap.get("summary") or "No prior meeting found.").strip()
        lines.extend([summary, ""])
        for key, heading in (
            ("follow_ups", "Open follow-ups"),
            ("decisions", "Decisions"),
            ("talking_points", "Suggested talking points"),
        ):
            section(lines, heading, recap.get(key), scopes)
    section(lines, "Agenda", agenda.get("items"), scopes)
    return "\n".join(lines).rstrip() + "\n"


def build_outputs(payload: dict, profile: dict) -> dict[Path, str | dict]:
    payload = normalize_payload(payload)
    if not payload.get("gtd_link"):
        payload["gtd_link"] = configured_gtd_link(profile)
    paths = {key: Path(value) for key, value in resolved_artifact_paths(profile["artifacts"]).items()}
    scopes = {scope["id"]: scope["name"] for scope in profile.get("scopes") or []}
    close_date = iso_date(payload.get("date") or date.today().isoformat(), "date")
    target = iso_date(payload.get("target_date") or close_date, "target_date")
    outputs: dict[Path, str | dict] = {
        paths["logs"] / f"EOD {close_date}.md": eod_markdown(payload, scopes),
        paths["plans"] / f"Daily Plan {target}.md": plan_markdown(payload, scopes),
        paths["tasks"] / f"Tasks {target}.md": task_markdown(payload, scopes),
        paths["state"] / f"{close_date}-close.json": payload,
    }
    for agenda in payload.get("agendas") or []:
        if not isinstance(agenda, dict):
            continue
        title = clean_filename(str(agenda.get("title") or "Meeting Agenda"))
        outputs[paths["agendas"] / target / f"{title}.md"] = agenda_markdown(agenda, scopes)
    return outputs


def export_paths(payload: dict, profile: dict) -> dict[str, list[tuple[Path, dict, str]] | Path]:
    payload = normalize_payload(payload)
    if not payload.get("gtd_link"):
        payload["gtd_link"] = configured_gtd_link(profile)
    paths = {key: Path(value) for key, value in resolved_artifact_paths(profile["artifacts"]).items()}
    close_date = iso_date(payload.get("date") or date.today().isoformat(), "date")
    target = iso_date(payload.get("target_date") or close_date, "target_date")
    exports = (profile.get("artifacts") or {}).get("exports") or {}
    scope_names = {scope["id"]: scope["name"] for scope in profile.get("scopes") or []}

    def labeled(values: Any) -> list[str]:
        return [label(value, scope_names) for value in (values or []) if label(value, scope_names).strip()]

    def labeled_meetings(values: Any) -> list[object]:
        rendered = []
        for value in values or []:
            if not isinstance(value, dict):
                rendered.append(value)
                continue
            item = copy.deepcopy(value)
            key = "title" if item.get("title") else "subject" if item.get("subject") else "title"
            item[key] = label({"text": item.get(key) or "Meeting", "scope_id": item.get("scope_id")}, scope_names)
            rendered.append(item)
        return rendered

    agenda_exports = []
    for agenda in payload.get("agendas") or []:
        if not isinstance(agenda, dict):
            continue
        item = copy.deepcopy(agenda)
        item["title"] = label({"text": item.get("title") or "Meeting Agenda", "scope_id": item.get("scope_id")}, scope_names)
        agenda_exports.append(item)
    result: dict[str, list[tuple[Path, dict, str]] | Path] = {"docx": []}
    if export_enabled(exports, "daily_plan_docx"):
        features = profile.get("features") or {}
        sections = payload.get("sections") or {}
        if payload.get("daily_plan"):
            daily_plan = copy.deepcopy(payload["daily_plan"])
        else:
            takeaway_value = payload.get("takeaways") or {}
            takeaway_limit = int(((features.get("daily_takeaways") or {}).get("max_items") or 3))
            daily_plan = {
                "date": target,
                "summary": payload.get("summary"),
                "meeting_insights": labeled(sections.get("meeting_insights")),
                "gtd_link": copy.deepcopy(payload.get("gtd_link")),
                "takeaways": {
                    "source_day": takeaway_value.get("source_day") or close_date,
                    "well": labeled(takeaway_value.get("well"))[:takeaway_limit],
                    "improve": labeled(takeaway_value.get("improve"))[:takeaway_limit],
                    "required_items": int(
                        ((features.get("daily_takeaways") or {}).get("required_items") or 0)
                    ),
                },
                "daily_big_3": labeled(sections.get("priorities")),
                "top_actions": labeled(sections.get("tasks")),
                "other_actions": labeled(sections.get("waiting")),
                "people_outreach": labeled(sections.get("people_outreach")),
                "meetings": labeled_meetings(sections.get("meetings")),
            }
        daily_plan["page_numbers"] = bool(features.get("docx_page_numbers", True))
        daily_plan.setdefault("takeaways", {}).setdefault(
            "required_items",
            int(((features.get("daily_takeaways") or {}).get("required_items") or 0)),
        )
        result["docx"].append((paths["plans"] / f"Daily Plan {target}.docx", daily_plan, "plan"))
    if export_enabled(exports, "agenda_docx"):
        for agenda in agenda_exports:
            if not isinstance(agenda, dict):
                continue
            title = clean_filename(str(agenda.get("title") or "Meeting Agenda"))
            result["docx"].append((paths["agendas"] / target / f"{title}.docx", agenda, "agenda"))
    if exports.get("xlsx"):
        result["xlsx"] = paths["tasks"] / f"Tasks {target}.xlsx"
    return result


def create_task_xlsx(payload: dict, profile: dict, output: Path) -> None:
    payload = normalize_payload(payload)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("XLSX export requires openpyxl") from exc
    scopes = {scope["id"]: scope["name"] for scope in profile.get("scopes") or []}
    tasks = ((payload.get("sections") or {}).get("tasks") or []) + ((payload.get("sections") or {}).get("carried") or [])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet.append(["Scope", "Task", "Status", "Due Date", "Source"])
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for item in tasks:
        if isinstance(item, dict):
            scope_id = item.get("scope_id")
            source = item.get("source") or ""
            if isinstance(source, dict):
                source = ":".join(
                    str(value) for value in (source.get("provider"), source.get("id")) if value
                )
            sheet.append([
                scopes.get(scope_id, scope_id or ""),
                item.get("text") or item.get("title") or "",
                item.get("status") or "Not Started",
                item.get("due_date") or "",
                source,
            ])
        else:
            sheet.append(["", str(item), "Not Started", "", ""])
    sheet.freeze_panes = "A2"
    for column, width in {"A": 22, "B": 60, "C": 18, "D": 16, "E": 30}.items():
        sheet.column_dimensions[column].width = width
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Create portable close-day artifacts.")
    parser.add_argument("--input", required=True, help="Approved close payload JSON.")
    parser.add_argument("--profile", help="Profile id from the local registry.")
    parser.add_argument("--profile-file", help="Explicit schema-v2 profile JSON path.")
    parser.add_argument("--config-root", help="Override the private config directory.")
    parser.add_argument("--approved", action="store_true", help="Required to write artifacts.")
    parser.add_argument("--dry-run", action="store_true", help="Print proposed files without writing.")
    args = parser.parse_args()

    if args.profile_file:
        profile = load_json(Path(args.profile_file).expanduser())
    else:
        profile, _ = resolve_profile(args.profile, Path(args.config_root).expanduser() if args.config_root else None)
    errors, _ = validate_profile(profile)
    if errors:
        raise ValueError("invalid profile: " + "; ".join(errors))
    payload = load_json(Path(args.input).expanduser())
    validate_required_takeaways(payload, profile)
    outputs = build_outputs(payload, profile)
    exports = export_paths(payload, profile)
    export_files = [path for path, _, _ in exports.get("docx", [])]
    if exports.get("xlsx"):
        export_files.append(exports["xlsx"])
    if args.dry_run:
        print(json.dumps({"files": [str(path) for path in [*outputs, *export_files]]}, indent=2))
        return 0
    if not args.approved:
        parser.error("artifact writes require --approved (use --dry-run to preview)")
    if not (profile.get("permissions") or {}).get("local_artifact_writes_enabled", False):
        parser.error("profile does not permit local artifact writes")
    existing = [path for path in [*outputs, *export_files] if path.exists()]
    if existing:
        raise FileExistsError(
            "refusing to overwrite existing artifact(s): " + ", ".join(str(path) for path in existing)
        )
    for path, content in outputs.items():
        path.parent.mkdir(parents=True, exist_ok=True)
        if isinstance(content, dict):
            atomic_write_json(path, content)
        else:
            atomic_write_text(path, content)
    for path, data, kind in exports.get("docx", []):
        if kind == "agenda":
            create_agenda_docx(data, path)
        else:
            create_daily_plan_docx(data, path)
    if exports.get("xlsx"):
        create_task_xlsx(payload, profile, exports["xlsx"])
    print(json.dumps({"written": [str(path) for path in [*outputs, *export_files]]}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
