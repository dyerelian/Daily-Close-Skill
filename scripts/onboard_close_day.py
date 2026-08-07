#!/usr/bin/env python3
"""Onboard a close-day installation from guided answers.

The LLM-facing part is intentionally prompt/answer based: `questions` emits the
question catalog and answer schema for Codex or another LLM to use, while `run`
performs deterministic file creation from the resulting answers JSON.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
MODULE_DIR = ROOT / "modules"
DEFAULT_REPORT_DIR = ROOT / "outputs" / "onboarding"

MODULE_ORDER = [
    "gmail-sweep",
    "calendar-outlook",
    "sent-mail-outlook",
    "teams-local-cache",
    "gtd-workbook",
    "daily-plan-docx",
    "crm-google-sheet",
    "granola-meetings",
    "slack-sweep",
    "source-of-truth",
]


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def resolve(path_text: str | None, base: Path = ROOT) -> Path | None:
    if not path_text:
        return None
    path = Path(path_text).expanduser()
    return path if path.is_absolute() else base / path


def rel_or_abs(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def load_manifests() -> dict[str, dict]:
    manifests: dict[str, dict] = {}
    for path in sorted(MODULE_DIR.glob("*.json")):
        manifest = load_json(path)
        manifests[manifest["id"]] = manifest
    return manifests


def default_answers() -> dict:
    return load_json(ROOT / "config" / "onboarding.answers.example.json")


def yes_no(prompt: str, default: bool = False) -> bool:
    suffix = "Y/n" if default else "y/N"
    raw = input(f"{prompt} [{suffix}]: ").strip().lower()
    if not raw:
        return default
    return raw in {"y", "yes", "true", "1"}


def ask(prompt: str, default: str = "") -> str:
    raw = input(f"{prompt}{f' [{default}]' if default else ''}: ").strip()
    return raw or default


def collect_interactive_answers() -> dict:
    answers = default_answers()
    owner = answers["owner"]
    systems = answers["systems"]

    owner["name"] = ask("Your name", owner["name"])
    owner["primary_email"] = ask("Primary email", owner["primary_email"])
    owner["timezone"] = ask("Timezone", owner["timezone"])
    owner["close_out_time"] = ask("Usual close-out time", owner["close_out_time"])

    email_provider = ask("Email provider (gmail/outlook/both/none)", "gmail").lower()
    systems["email"]["provider"] = email_provider
    systems["email"]["account"] = owner["primary_email"]

    systems["calendar"]["provider"] = ask("Calendar provider (outlook/none)", "outlook").lower()
    systems["tasks"]["mode"] = ask("Tasks mode (create_new/existing/skip)", "create_new").lower()
    if systems["tasks"]["mode"] == "existing":
        systems["tasks"]["existing_path"] = ask("Existing GTD workbook path")
    systems["tasks"]["allow_writes_after_approval"] = yes_no(
        "Allow task workbook writes after proposal approval?", False
    )

    systems["crm"]["mode"] = ask("CRM mode (create_new/existing/skip)", "create_new").lower()
    if systems["crm"]["mode"] == "existing":
        systems["crm"]["existing_path"] = ask("Existing CRM workbook or Sheet export path")
    systems["crm"]["allow_live_sheet_writes"] = yes_no(
        "Allow live CRM sheet writes after proposal approval?", False
    )

    systems["documents"]["daily_plan"] = ask("Daily plan output (docx/markdown/skip)", "docx").lower()
    systems["meetings"]["granola"] = yes_no("Use Granola meeting notes?", False)
    systems["meetings"]["teams_local_cache"] = yes_no("Use Teams local cache?", False)
    systems["chat"]["slack"] = yes_no("Use Slack sweep?", False)
    systems["source_of_truth"]["mode"] = ask(
        "Source-of-truth mode (local_files/confluence/google_docs/none)", "local_files"
    ).lower()
    answers["runtime"]["artifact_root"] = ask(
        "Local artifact root", answers["runtime"]["artifact_root"]
    )
    excluded = ask("Topics/accounts/projects to exclude (comma-separated)")
    answers["scope_exclusions"] = {
        "match_mode": "case_insensitive_term",
        "topics": [
            {
                "name": term,
                "match_terms": [term],
                "reason": "User-requested exclusion during onboarding",
            }
            for term in (item.strip() for item in excluded.split(","))
            if term
        ],
    }
    return answers


def connector_configured(answers: dict, connector: str) -> bool:
    value = ((answers.get("connectors") or {}).get(connector) or {}).get("configured")
    return bool(value)


def build_seed_queries(answers: dict) -> list[dict]:
    email = answers["systems"]["email"]
    default_window = "newer_than:365d"
    queries = []
    for seed in answers["systems"]["crm"].get("seed_terms") or []:
        terms = [str(term).strip() for term in seed.get("query_terms") or [] if str(term).strip()]
        if not terms:
            continue
        query_terms = " OR ".join(f'"{term}"' if " " in term else term for term in terms)
        queries.append(
            {
                "category": seed.get("category") or "partner_led_pipeline",
                "account_hint": seed.get("account_hint") or terms[0],
                "query": f"{default_window} -in:spam -in:trash ({query_terms})",
            }
        )
    if not queries and email.get("provider") in {"gmail", "both"}:
        queries = load_json(ROOT / "config" / "daily-close.example.json")["crm"]["gmail_seed_queries"]
    return queries


def selected_modules(answers: dict) -> list[str]:
    systems = answers["systems"]
    modules: list[str] = []
    email_provider = (systems["email"].get("provider") or "none").lower()
    calendar_provider = (systems["calendar"].get("provider") or "none").lower()

    if email_provider in {"gmail", "both"}:
        modules.append("gmail-sweep")
    if calendar_provider in {"outlook", "both"}:
        modules.append("calendar-outlook")
    if email_provider in {"outlook", "both"}:
        modules.append("sent-mail-outlook")
    if systems["meetings"].get("teams_local_cache"):
        modules.append("teams-local-cache")
    if systems["tasks"].get("mode") not in {"skip", "none"}:
        modules.append("gtd-workbook")
    if systems["documents"].get("daily_plan") not in {"skip", "none"}:
        modules.append("daily-plan-docx")
    if systems["crm"].get("mode") not in {"skip", "none"}:
        modules.append("crm-google-sheet")
    if systems["meetings"].get("granola"):
        modules.append("granola-meetings")
    if systems["chat"].get("slack"):
        modules.append("slack-sweep")
    if systems["source_of_truth"].get("mode") not in {"skip", "none"}:
        modules.append("source-of-truth")

    ordered = [module for module in MODULE_ORDER if module in set(modules)]
    return ordered


def build_config(answers: dict, artifact_root: Path) -> dict:
    owner = answers["owner"]
    systems = answers["systems"]
    write_policy = answers.get("write_policy") or {}
    modules = selected_modules(answers)

    gtd_path = systems["tasks"].get("existing_path")
    if systems["tasks"].get("mode") == "create_new":
        gtd_path = str(artifact_root / "GTD" / "close-day-gtd.xlsx")

    crm_path = systems["crm"].get("existing_path")
    if systems["crm"].get("mode") == "create_new":
        crm_path = str(artifact_root / "CRM" / "close-day-crm.xlsx")

    daily_plan_dir = systems["documents"].get("daily_plan_dir") or str(artifact_root / "Daily Plan")
    eod_log_dir = systems["documents"].get("eod_log_dir") or str(Path(daily_plan_dir) / "GTD Daily Logs")
    agenda_dir = systems["documents"].get("agenda_dir") or str(Path(daily_plan_dir) / "Agendas")

    gmail_account = systems["email"].get("account") or owner["primary_email"]
    routine_query = systems["email"].get("routine_query") or "newer_than:7d -in:spam -in:trash"

    return {
        "profile_name": f"{owner['name'].strip().lower().replace(' ', '-')}-close-day",
        "schema_version": 1,
        "owner": {
            "name": owner["name"],
            "primary_email": owner["primary_email"],
            "timezone": owner["timezone"],
            "workdays": owner.get("workdays") or [],
            "close_out_time": owner.get("close_out_time"),
        },
        "privacy": answers.get("privacy") or {},
        "scope_exclusions": answers.get("scope_exclusions")
        or {"match_mode": "case_insensitive_term", "topics": []},
        "write_mode": {
            "enabled": bool(write_policy.get("enabled")),
            "require_single_approval": True,
            "external_writes_enabled": bool(write_policy.get("external_writes_enabled")),
            "document_generation_enabled": bool(write_policy.get("document_generation_enabled")),
            "crm_writes_enabled": bool(write_policy.get("crm_writes_enabled")),
        },
        "enabled_modules": modules,
        "paths": {
            "python": "python",
            "gtd_workbook": gtd_path,
            "daily_plan_dir": daily_plan_dir,
            "eod_log_dir": eod_log_dir,
            "agenda_dir": agenda_dir,
            "add_gtd_items_script": None,
            "agenda_creator_script": None,
            "source_of_truth_skill": None,
            "source_of_truth_map": str(artifact_root / "Source of Truth" / "source-of-truth-map.csv"),
        },
        "connectors": {
            "gmail": {
                "account": gmail_account,
                "configured": connector_configured(answers, "gmail"),
                "max_results_per_query": 10,
                "default_window": "newer_than:365d",
            },
            "slack": {
                "configured": connector_configured(answers, "slack"),
                "workspace": None,
            },
            "granola": {
                "configured": connector_configured(answers, "granola"),
            },
            "atlassian": {
                "configured": connector_configured(answers, "atlassian"),
            },
        },
        "crm": {
            "workbook_path": crm_path,
            "csv_seed_dir": str(artifact_root / "CRM" / "csv_seed"),
            "proposal_output_dir": str(artifact_root / "CRM" / "proposals"),
            "gmail_seed_queries": build_seed_queries(answers),
        },
        "modules": {
            "calendar-outlook": {
                "enabled": "calendar-outlook" in modules,
                "days_ahead": int(systems["calendar"].get("days_ahead") or 1),
            },
            "sent-mail-outlook": {
                "enabled": "sent-mail-outlook" in modules,
                "window": "today",
            },
            "teams-local-cache": {
                "enabled": "teams-local-cache" in modules,
                "window": "today",
                "cache_path": None,
            },
            "gtd-workbook": {
                "enabled": "gtd-workbook" in modules,
                "workbook_path": gtd_path,
                "allow_writes": bool(systems["tasks"].get("allow_writes_after_approval")),
            },
            "daily-plan-docx": {
                "enabled": "daily-plan-docx" in modules,
                "output_dir": daily_plan_dir,
                "allow_generation": bool(write_policy.get("document_generation_enabled")),
            },
            "source-of-truth": {
                "enabled": "source-of-truth" in modules,
                "mode": systems["source_of_truth"].get("mode"),
                "allow_confluence_writes": bool(systems["source_of_truth"].get("allow_confluence_writes")),
            },
            "gmail-sweep": {
                "enabled": "gmail-sweep" in modules,
                "account": gmail_account,
                "default_query": routine_query,
            },
            "granola-meetings": {
                "enabled": "granola-meetings" in modules,
                "window": "today",
            },
            "slack-sweep": {
                "enabled": "slack-sweep" in modules,
                "window": "today",
                "channels": systems["chat"].get("slack_channels") or [],
            },
            "crm-google-sheet": {
                "enabled": "crm-google-sheet" in modules,
                "workbook_path": crm_path,
                "proposal_output_dir": str(artifact_root / "CRM" / "proposals"),
                "allow_live_sheet_writes": bool(systems["crm"].get("allow_live_sheet_writes")),
            },
        },
    }


def create_gtd_workbook(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    sheets = {
        "Next Actions": [
            "ID",
            "Action",
            "Project",
            "Context",
            "Priority",
            "Status",
            "Due Date",
            "Scheduled Date",
            "Notes",
        ],
        "Waiting For": [
            "ID",
            "Who",
            "What",
            "Project",
            "Status",
            "Requested Date",
            "Follow-up Date",
            "Notes",
        ],
        "Inbox": [
            "ID",
            "Captured",
            "Item",
            "Source",
            "Processed?",
            "Decision",
            "Notes",
        ],
        "Projects": [
            "ID",
            "Project",
            "Status",
            "Area",
            "Owner",
            "Last Reviewed",
            "Notes",
        ],
        "Lists": [
            "Statuses",
            "Priorities",
            "Contexts",
            "Decisions",
        ],
    }
    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for sheet_name, headers in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.append(headers)
        ws.append(["" for _ in headers])
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[chr(64 + idx)].width = max(14, min(36, len(header) + 8))
        if sheet_name != "Lists":
            ref = f"A1:{chr(64 + len(headers))}2"
            table = Table(displayName=sheet_name.replace(" ", "") + "Table", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(table)
    lists = wb["Lists"]
    values = {
        "A": ["Not Started", "In Progress", "Waiting", "Complete", "Canceled"],
        "B": ["P1 - Must", "P2 - Should", "P3 - Could", "P4 - Later"],
        "C": ["Computer", "Email", "Call", "Errand", "Meeting"],
        "D": ["Do", "Delegate", "Defer", "Delete", "Clarify", "Incubate"],
    }
    for col, rows in values.items():
        for idx, value in enumerate(rows, start=2):
            lists[f"{col}{idx}"] = value
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def create_source_of_truth_files(root: Path) -> list[str]:
    folder = root / "Source of Truth"
    folder.mkdir(parents=True, exist_ok=True)
    mapping = folder / "source-of-truth-map.csv"
    if not mapping.exists():
        with mapping.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Type", "Name", "Canonical URL", "Status", "Last Reviewed", "Notes"])
    template = folder / "canonical-page-template.md"
    if not template.exists():
        template.write_text(
            "# {{Project or Account Name}}\n\n"
            "## Overview\n\n"
            "## Current Status\n\n"
            "## Decisions\n\n"
            "## Open Questions\n\n"
            "## Updates\n\n",
            encoding="utf-8",
        )
    return [str(mapping), str(template)]


def create_daily_plan_dirs(config: dict) -> list[str]:
    created = []
    for key in ("daily_plan_dir", "eod_log_dir", "agenda_dir"):
        value = config["paths"].get(key)
        if not value:
            continue
        path = Path(value)
        path.mkdir(parents=True, exist_ok=True)
        created.append(str(path))
    settings = Path(config["paths"]["daily_plan_dir"]) / "daily-plan-settings.json"
    if not settings.exists():
        write_json(
            settings,
            {
                "quote_rotation": "avoid_recent_repeats",
                "first_task_policy": "send_next_workday_agendas_when_available",
                "created_by": "close-day onboarding",
            },
        )
    created.append(str(settings))
    return created


def create_crm_artifacts(config: dict) -> list[str]:
    sys.path.insert(0, str(ROOT / "scripts"))
    from generate_crm_workbook import create_workbook

    workbook = Path(config["crm"]["workbook_path"])
    csv_dir = Path(config["crm"]["csv_seed_dir"])
    create_workbook(workbook, csv_dir)
    Path(config["crm"]["proposal_output_dir"]).mkdir(parents=True, exist_ok=True)
    return [str(workbook), str(csv_dir), config["crm"]["proposal_output_dir"]]


def create_artifacts(config: dict, answers: dict) -> tuple[list[str], list[str]]:
    created: list[str] = []
    skipped: list[str] = []
    systems = answers["systems"]

    if "gtd-workbook" in config["enabled_modules"]:
        if systems["tasks"].get("mode") == "create_new":
            create_gtd_workbook(Path(config["paths"]["gtd_workbook"]))
            created.append(config["paths"]["gtd_workbook"])
        else:
            skipped.append("GTD workbook creation skipped; existing task system selected.")

    if "crm-google-sheet" in config["enabled_modules"]:
        if systems["crm"].get("mode") == "create_new":
            created.extend(create_crm_artifacts(config))
        else:
            skipped.append("CRM workbook creation skipped; existing CRM selected.")

    if "daily-plan-docx" in config["enabled_modules"]:
        created.extend(create_daily_plan_dirs(config))

    if "source-of-truth" in config["enabled_modules"]:
        mode = systems["source_of_truth"].get("mode")
        if mode == "local_files":
            artifact_root = Path(config["paths"]["source_of_truth_map"]).parents[0].parents[0]
            created.extend(create_source_of_truth_files(artifact_root))
        else:
            skipped.append(f"Source-of-truth local files skipped; mode is {mode}.")

    return created, skipped


def validate_setup(config: dict, manifests: dict[str, dict]) -> dict:
    blockers: list[str] = []
    gaps: list[str] = []
    notes: list[str] = []
    enabled = config.get("enabled_modules") or []

    if not enabled:
        blockers.append("No modules are enabled.")

    connectors = config.get("connectors") or {}
    for module_id in enabled:
        manifest = manifests.get(module_id)
        if not manifest:
            blockers.append(f"Unknown enabled module: {module_id}")
            continue
        for connector in manifest.get("required_connectors") or []:
            if str(connector).startswith("local:"):
                continue
            if (
                module_id == "source-of-truth"
                and connector == "atlassian"
                and ((config.get("modules") or {}).get("source-of-truth") or {}).get("mode")
                != "confluence"
            ):
                continue
            configured = bool((connectors.get(connector) or {}).get("configured", True))
            if connector not in connectors or not configured:
                gaps.append(f"{module_id}: connector {connector} is not marked configured.")

    path_requirements = {
        "gtd-workbook": [("paths.gtd_workbook", config["paths"].get("gtd_workbook"))],
        "daily-plan-docx": [
            ("paths.daily_plan_dir", config["paths"].get("daily_plan_dir")),
            ("paths.eod_log_dir", config["paths"].get("eod_log_dir")),
        ],
        "crm-google-sheet": [("crm.workbook_path", config["crm"].get("workbook_path"))],
        "source-of-truth": [
            ("paths.source_of_truth_map", config["paths"].get("source_of_truth_map"))
        ],
    }
    for module_id, checks in path_requirements.items():
        if module_id not in enabled:
            continue
        for label, value in checks:
            if not value:
                gaps.append(f"{module_id}: {label} is not set.")
                continue
            if not Path(value).exists():
                gaps.append(f"{module_id}: {label} does not exist yet: {value}")

    if not config.get("write_mode", {}).get("enabled"):
        notes.append("write_mode.enabled is false; close-day will remain proposal/local-template only.")
    status = "ready"
    if blockers:
        status = "blocked"
    elif gaps:
        status = "usable_with_gaps"
    return {
        "status": status,
        "blockers": blockers,
        "gaps": gaps,
        "notes": notes,
        "enabled_modules": enabled,
    }


def write_report(path: Path, answers: dict, config: dict, validation: dict, created: list[str], skipped: list[str]) -> None:
    lines = [
        "# close-day onboarding report",
        "",
        f"Generated: {datetime.now(timezone.utc).replace(microsecond=0).isoformat()}",
        f"Status: {validation['status']}",
        "",
        "## Enabled modules",
        "",
    ]
    lines.extend(f"- {module}" for module in config.get("enabled_modules") or [])
    lines.extend(["", "## Created artifacts", ""])
    lines.extend(f"- {item}" for item in created) if created else lines.append("- None")
    lines.extend(["", "## Skipped artifacts", ""])
    lines.extend(f"- {item}" for item in skipped) if skipped else lines.append("- None")
    lines.extend(["", "## Gaps", ""])
    lines.extend(f"- {item}" for item in validation["gaps"]) if validation["gaps"] else lines.append("- None")
    lines.extend(["", "## Notes", ""])
    lines.extend(f"- {item}" for item in validation["notes"]) if validation["notes"] else lines.append("- None")
    lines.extend(
        [
            "",
            "## Privacy",
            "",
            f"- LLM data level: {answers.get('privacy', {}).get('llm_data_level')}",
            f"- Allow raw external content: {answers.get('privacy', {}).get('allow_raw_external_content')}",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def question_catalog() -> dict:
    manifests = load_manifests()
    return {
        "schema_version": 1,
        "answer_file": "config/onboarding.answers.example.json",
        "core_questions": [
            "Who is the owner: name, primary email, timezone, workdays, and close-out time?",
            "Is this a solo operator setup or team workspace?",
            "Which modules should be enabled: email, calendar, tasks, CRM, meetings, chat, documents, source-of-truth?",
            "Which artifacts already exist, and which should setup create locally?",
            "Which connectors are authenticated?",
            "Which write targets may be enabled after the single approval gate?",
            "What data may be sent to an LLM during onboarding and daily use?",
            "Which topics, accounts, or projects should every daily close exclude?",
        ],
        "modules": {
            module_id: manifest.get("onboarding", {})
            for module_id, manifest in manifests.items()
        },
    }


def llm_prompt() -> str:
    catalog = question_catalog()
    example = default_answers()
    return (
        "You are onboarding the close-day Codex skill for a new user.\n"
        "Ask concise questions in sections, choose modules based on answers, and return a JSON answer file.\n"
        "Do not request raw mailbox, Slack, calendar, or document contents during onboarding unless the user explicitly allows it.\n\n"
        "Return JSON matching this example shape:\n\n"
        f"```json\n{json.dumps(example, indent=2)}\n```\n\n"
        "Use these module-specific onboarding prompts:\n\n"
        f"```json\n{json.dumps(catalog['modules'], indent=2)}\n```\n"
    )


def run_command(args: argparse.Namespace) -> int:
    answers = load_json(resolve(args.answers)) if args.answers else collect_interactive_answers()
    artifact_root = resolve(answers["runtime"].get("artifact_root")) or (Path.home() / "Documents" / "close-day")
    if args.dry_run:
        artifact_root = DEFAULT_REPORT_DIR / "dry-run-workspace"
    config = build_config(answers, artifact_root)
    config_out = resolve(args.config_out) if args.config_out else ROOT / "config" / "daily-close.local.json"
    if args.dry_run:
        config_out = DEFAULT_REPORT_DIR / "dry-run-daily-close.local.json"

    created, skipped = create_artifacts(config, answers)
    write_json(config_out, config)

    validation = validate_setup(config, load_manifests())
    access_check = {
        **validation,
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "config": str(config_out),
    }
    report_dir = resolve(args.report_dir) or DEFAULT_REPORT_DIR
    write_json(report_dir / "access-check.json", access_check)
    write_report(report_dir / "setup-report.md", answers, config, validation, created, skipped)

    if args.json:
        print(json.dumps(access_check, indent=2))
    else:
        print(f"status: {validation['status']}")
        print(f"config: {config_out}")
        print(f"report: {report_dir / 'setup-report.md'}")
        print(f"access_check: {report_dir / 'access-check.json'}")
    return 1 if validation["status"] == "blocked" else 0


def validate_command(args: argparse.Namespace) -> int:
    config = load_json(resolve(args.config))
    validation = validate_setup(config, load_manifests())
    if args.json:
        print(json.dumps(validation, indent=2))
    else:
        print(f"status: {validation['status']}")
        for item in validation["blockers"]:
            print(f"blocker: {item}")
        for item in validation["gaps"]:
            print(f"gap: {item}")
        for item in validation["notes"]:
            print(f"note: {item}")
    return 1 if validation["status"] == "blocked" else 0


def questions_command(args: argparse.Namespace) -> int:
    payload = question_catalog() if args.json else llm_prompt()
    if args.out:
        out = resolve(args.out)
        out.parent.mkdir(parents=True, exist_ok=True)
        if args.json:
            write_json(out, payload)
        else:
            out.write_text(payload, encoding="utf-8")
        print(f"wrote: {out}")
    else:
        print(json.dumps(payload, indent=2) if args.json else payload)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Onboard a close-day skill installation.")
    sub = parser.add_subparsers(dest="command", required=True)

    q = sub.add_parser("questions", help="Emit LLM/Codex onboarding questions and answer schema.")
    q.add_argument("--out", help="Optional output prompt/catalog path.")
    q.add_argument("--json", action="store_true", help="Emit machine-readable question catalog.")
    q.set_defaults(func=questions_command)

    run = sub.add_parser("run", help="Create config and local setup artifacts from answers.")
    run.add_argument("--answers", help="Answers JSON. If omitted, ask in the console.")
    run.add_argument("--config-out", help="Config output path. Defaults to config/daily-close.local.json.")
    run.add_argument("--report-dir", help="Report output directory. Defaults to outputs/onboarding.")
    run.add_argument("--dry-run", action="store_true", help="Write generated files under outputs/onboarding.")
    run.add_argument("--json", action="store_true", help="Emit access-check JSON.")
    run.set_defaults(func=run_command)

    val = sub.add_parser("validate", help="Validate setup readiness from a generated config.")
    val.add_argument("--config", default="config/daily-close.local.json")
    val.add_argument("--json", action="store_true")
    val.set_defaults(func=validate_command)

    args = parser.parse_args()
    try:
        return args.func(args)
    except KeyboardInterrupt:
        print("canceled", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
